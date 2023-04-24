import requests
from bs4 import BeautifulSoup
import openpyxl

url = 'https://nj.lianjia.com/ershoufang/103124821277.html'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36 Edge/16.16299'
}
response = requests.get(url, headers=headers)
soup = BeautifulSoup(response.text, 'html.parser')

# 获取房屋信息
house_info = soup.find('div', class_='houseInfo').text.strip()

# 获取小区信息
community_info = soup.find('div', class_='communityName').find('a').text.strip()

# 获取房屋总价
total_price = soup.find('span', class_='total').text.strip()

# 获取房屋单价
unit_price = soup.find('span', class_='unitPriceValue').text.strip()

# 获取房屋基本信息
basic_info_list = soup.find_all('div', class_='baseInfoItem')
basic_info_dict = {}
for basic_info in basic_info_list:
    key = basic_info.find('span', class_='name').text.strip()
    value = basic_info.find('span', class_='content').text.strip()
    basic_info_dict[key] = value

# 获取交易属性信息
transaction_list = soup.find_all('div', class_='transactionItem')
transaction_dict = {}
for transaction in transaction_list:
    key = transaction.find('span', class_='label').text.strip()
    value = transaction.find('span', class_='content').text.strip()
    transaction_dict[key] = value

# 将获取到的信息存入Excel表格
wb = openpyxl.Workbook()
ws = wb.active

ws['A1'] = '房屋信息'
ws['B1'] = '小区信息'
ws['C1'] = '房屋总价'
ws['D1'] = '房屋单价'

ws['A2'] = house_info
ws['B2'] = community_info
ws['C2'] = total_price
ws['D2'] = unit_price

for i, (key, value) in enumerate(basic_info_dict.items()):
    ws.cell(row=1, column=i+5).value = key
    ws.cell(row=2, column=i+5).value = value

for i, (key, value) in enumerate(transaction_dict.items()):
    ws.cell(row=1, column=i+5+len(basic_info_dict)).value = key
    ws.cell(row=2, column=i+5+len(basic_info_dict)).value = value

wb.save('house_info.xlsx')